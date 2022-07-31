from pynput.keyboard import Key as keys,Controller as ControllerK#key
import pyperclip as pyp#key
import pyautogui as pya#press,key
import os#press
from datetime import datetime,timedelta#excel
import openpyxl,xlrd#excel
import pyodbc,mysql.connector#DB
##Threading##
import sys, trace, threading
import decimal

class press(object):
    def __init__(self,*args,mode='and'):
        completa=lambda x:x if '.png' in x.lower() else x+'.png'
        self.__root="\\".join(os.getcwd().split('\\')+['Imagenes'])
        self.image=[completa(x) for x in args]
        self.acciones={'L':'left','R':'right','M':'middle'}
        self.mode=mode.lower()

    def clickh(self,x0=0,y0=0,x1=0,y1=0,mensaje=True,threshold=0.9,grayscale=False,duration=0,**kwargs):
        completa=lambda x:x if '.png' in x.lower() else x+'.png'
        if kwargs!={}:self.image=[completa(img)]
        pya.FAILSAFE = False #bloquea el error que se muestra cuando se mueve el mouse en plena acción de la función click
        for img in self.image:
            image_pos = pya.locateOnScreen(os.path.join(self.__root,img), confidence=threshold, grayscale=grayscale)
            if isinstance(image_pos,pya.pyscreeze.Box):
                if mensaje:print("Se encontró la imagen...",img.split('\\')[-1])
                xx,yy=[pya.center(image_pos)._asdict()[x] for x in 'xy']
                pya.moveTo(x=xx+x0,y=yy+y0)
                pya.drag(xOffset=x1,yOffset=y1,duration=duration)
                return
            if mensaje:print("No se ha encontrado imagen...",img.split('\\')[-1])
    
    def click(self,x=0,y=0,clks=1,accion='L',step=0,threshold=0.9,duration=0,mensaje=True,grayscale=False,**kwargs):
        completa=lambda x:x if '.png' in x.lower() else x+'.png'
        if kwargs!={}:self.image=[completa(img)]
        pya.FAILSAFE = False #bloquea el error que se muestra cuando se mueve el mouse en plena acción de la función click
        for img in self.image:
            image_pos = pya.locateOnScreen(os.path.join(self.__root,img), confidence=threshold, grayscale=grayscale)
            if isinstance(image_pos,pya.pyscreeze.Box):
                if mensaje:print("Se encontró la imagen...",img.split('\\')[-1])
                xx,yy=[pya.center(image_pos)._asdict()[x] for x in 'xy']
                pya.click(button=self.acciones[accion],x=xx+x,y=yy+y,clicks=clks,interval=step,duration=duration)
                return
            if mensaje:print("No se ha encontrado imagen...",img.split('\\')[-1])
                
    def isimage(self,*isnot,mode='or|or',threshold=0.9,mensaje=True,grayscale=False):
        """Si mode='or'  se forma (is1 or is2 or ...) or (notis1 or notis2 or ...)
           Si mode='and' se forma (is1 or is2 or ...) and (notis1 or notis2 or ...)"""
        mode=mode.split('|')
        completa=lambda x:x if '.png' in x.lower() else x+'.png'
        isnot=[completa(x) for x in isnot]
        self.__is=False
        for img in self.image:
            image_pos = pya.locateOnScreen(os.path.join(self.__root,img), confidence=threshold, grayscale=grayscale)
            if isinstance(image_pos,pya.pyscreeze.Box):
                self.__is=True
                if mode[0]=='or':break
            else:
                if mode[0]=='and':
                    self.__is=False
                    break
        if isnot==[]:
            if mensaje:print(f"{'No'*int(not(self.__is))} se encontró...",self.image)
            return self.__is
        self.__notis=False
        for img in isnot:
            image_pos = pya.locateOnScreen(os.path.join(self.__root,img), confidence=threshold, grayscale=grayscale)
            if image_pos==None:
                self.__notis=True
                if mode[1]=='or':break
            else:
                if mode[1]=='and':
                    self.__notis=False
                    break
        if self.image==[]:
            if mensaje:print(f"{'No'*int(self.__notis)} se encontró...",self.image)
            return self.__notis
        if mensaje:print(f"{'No'*int(not(self.__is and self.__notis))} se encontró...",self.image)
        if self.mode=='or':return self.__is or self.__notis
        return self.__is and self.__notis
    
    def wclick(self,x=0,y=0,clks=1,accion='L',step=0,threshold=0.9,duration=0,mensaje=True,time=0.3,maxi=1000):
        """La Funcion espera encontrar la imagen para luego hacerle click
           la cantidad de intentos máximo por default es maxi=1000"""
        count=0
        while True:
            pya.sleep(time)
            count+=1
            if press(*self.image).isimage():press(*self.image).click(clks=clks);break
            if count==maxi:break

    def clickw(self,*wait,time=0.3,maxi=1000,mode='or|or',**kwargs):
        """Wait es la imagen con la que va a esperar
           pone en la imagen el simbolo ! para diferenciarlo, ejemplo
           clickw('some.png','!some2.png')
        """
        completa=lambda x:x if '.png' in x.lower() else x+'.png'
        isim=[completa(x) for x in wait if '!' not in x]
        notisim=[completa(x.replace('!','')) for x in wait if completa(x) not in isim]
        count=0
        while True:
            press(*self.image).click(**kwargs)
            pya.sleep(time)
            count+=1
            if press(*isim,mode=self.mode).isimage(*notisim,mode=mode):break
            if count==maxi:break

    

    def cwrite(self,text='',select=True,time=0.2,maxi=1000,condi='True',**kwargs):
        """
           Función que escribe un speech, y valida que lo ha escrito, si el resultado es positivo rompe lazo.
           
           text(''): <string>,escribe el contenido y lo copia para validar que lo ha escrito, 
           select(True): <bool>, indica si selecciona el texto con (ctrl+a) antes.
           time(0.2): <float>, indica el tiempo(en segundos) que espera entre el click y la copia del texto
           maxi(1000): <int>, indica la máxima cantidad de intentos de copia del texto antes de salirse del lazo

           Estructura:

               press(*args,mode='and').cwrite(text=',select=True',time=0.2,maxi=1000,**kwargs)
        """
        count=0
        while True:
            press(*self.image).click(**kwargs)
            pya.sleep(time)
            key(str(text)).Write(select=select)
            pya.sleep(time)
            val=key().copy(select=select)
            count+=1
            if str(text)==val:
                if self.mode=='and' and eval(condi):break
            elif self.mode=='or' and eval(condi):break
            if count==maxi:break
            
    def ccopy(self,text='',select=True,time=0.2,maxi=1000,condi='True',**kwargs):
        """
           Función que hace click sobre un texto y copia  y retorna el texto siempre y cuando
           cumple las condicionales fijadas condicionales.
           
           condi(True): Es un <string> donde se pueden establecer condicionales adicionales a las existentes para romper el lazo
               Ejemplo: condi='val.isdigit() or val.isalpha()'
           text(''): Es un <string> que se iguala a lo que se copia y si es igual rompe lazo y retorna el string
           select(True): <bool>, indica si selecciona el texto con (ctrl+a) antes de copiarlo
           time(0.2): <float>, indica el tiempo(en segundos) que espera entre el click y la copia del texto
           maxi(1000): <int>, indica la máxima cantidad de intentos de copia del texto antes de salirse del lazo
           self.mode('and'):parametro de la clase que permite darle mayor flexibilidad al parametro 'condi', ya que si es 'or', este si puede retornar vacio ''

           Estructura:

               press(*args,mode='and').ccopy(text='',select=True,time=0.2,maxi=1000,condi='True',**kwargs)
        """
        count=0
        while True:
            press(*self.image).click(**kwargs)
            pya.sleep(time)
            val=key().copy(select=select)
            count+=1
            if val!=''and (text=='' or val==text):
                if self.mode=='and' and eval(condi):return val
            elif self.mode=='or' and eval(condi):return val
            if count==maxi:break

    def wait(self,*wait,time=0.3,maxi=1000):
        """
           Función que se queda en espera hasta que aparezca y/o desaparezca una imagen.
                      
           time(0.2): <float>, indica el tiempo(en segundos) que espera entre el click y la copia del texto
           maxi(1000): <int>, indica la máxima cantidad de intentos de copia del texto antes de salirse del lazo
           self.mode('and'):parametro que se usa expresar si es un "and" o un "or" para unir lo que espera hasta que aparezco de lo que desaparezca

           Estructura:

               press(*args,mode='and').wait(*wait,time=0.3,maxi=1000)
               
           Estructura funcional:
           
               press(wait until appear).wait(wait until disappear)
        """
        count=0
        while True:
            pya.sleep(time)
            count+=1
            if press(*self.image,mode=self.mode).isimage(*wait):break
            if count==maxi:break


class key(object):
    def __init__(self,text=''):
        self.text=text
        self.keyboard = ControllerK()# permite escribir caracteres especiales como ñ,á,é,etc

    def Write(self,text='',select=True):
        if select:pya.hotkey('ctrl', 'a')
        if text!='':self.text=text
        inter= [None]+[idx for idx,x in enumerate(self.text) if x in 'áéíóúÁÉÍÓÚñÑ+']+[None]
        if inter==[None,None]:
            pya.typewrite(self.text) 
            return
        for x in zip(inter[:-1],inter[1:]):
            if self.text[x[0]:x[1]]=='':continue
            if len(self.text[x[0]:x[1]])==1:self.__Buttons(self.text[x[0]:x[1]]);continue
            self.__Buttons(self.text[x[0]:x[1]][0])
            pya.typewrite(self.text[x[0]:x[1]][1:])

    def __Buttons(self,text):
        if not(self.keyboard):self.keyboard = ControllerK()
        self.keyboard.press(text)
        self.keyboard.release(text)

    def Arrow(self,key="d",num=1,time=0.3):
        select={"w":lambda:self.__Buttons(keys.up),"a":lambda:self.__Buttons(keys.left),"s":lambda:self.__Buttons(keys.down),"d":lambda:self.__Buttons(keys.right),
                't':lambda:self.__Buttons(keys.tab),'st':lambda:pya.hotkey('shift','tab')}#t:tab,st:shift+tab
        for i in range(num):
            select[key]()
            pya.sleep(time)

    def scut(self,*args):
        """shortcut"""
        pya.hotkey(*args)

    def scroll(self,x=0,y=-500):
        if y!=0:#up down
            pya.scroll(y)
        if x!=0:#right left
            pya.hscroll(x)
        
    def copy(self,men='',select=True):
        pyp.copy(men)
        if men!='':return
        if select:pya.hotkey('ctrl', 'a')
        pya.sleep(0.2)
        pya.hotkey('ctrl', 'c')
        pya.hotkey('ctrl', 'c')
        pya.sleep(0.3)
        Result = pyp.paste()
        print(Result)
        return Result

    def paste(self):
        pya.hotkey('ctrl','v')


class DB(object):
    def __init__(self,server,database,user,password):
        """server,database,user,password"""
        self.credenciales=[server,database,user,password]
    
    def SQL2Dict(self,Select,resultados=True):
        SQLServer = pyodbc.connect('DRIVER=%s;SERVER={};DATABASE={};UID={};PWD={}'.format(*self.credenciales)%('{SQL Server}'), autocommit=True)
        cursor = SQLServer.cursor()
        cursor.execute(Select)
        cursor.commit()
        if resultados:
            columnas = [column[0].upper() for column in cursor.description]
            resultado = [{y:row[idx] for idx,y in enumerate(columnas)} for row in cursor.fetchall()]
            return resultado
        
    def Mysql2Dict(self,Select,resultados=True):
        head=['host','database','user','password']
        crede=dict(zip(head,self.credenciales))
        cnxn = mysql.connector.connect(**crede)
        cursor = cnxn.cursor(buffered=True)
        cursor.execute(Select)
        cnxn.commit()
        if resultados:
            columnas = [column[0].upper() for column in cursor.description]
            resultado = [{y:row[idx] for idx,y in enumerate(columnas)} for row in cursor.fetchall()]
            return resultado
        
class Excel(object):
    def __init__(self,excel):
        self.excel=excel

    def Xlsx2Dict(self,excel='', sheet=1,read_only=False):
        """sheet=1 se escoge la primera pagina
           sheet='Hoja 1', se escoge la pagina con el nombre Hoja 1"""
        def filtro(value):
            value = value if value not in (None,'None','',' ') else ''
            value = value if not(isinstance(value, datetime)) else value.strftime("%Y-%m-%d")#"%d/%m/%Y")
            return str(value).replace("'", "")
        if excel!='':self.excel=excel
        _Writer = openpyxl.load_workbook(filename=self.excel, read_only=read_only)
        sheets = _Writer.sheetnames
        if isinstance(sheet,str):_Writer.active=sheets.index(sheet)
        else:_Writer.active=sheet-1#en esta funcion se considera que la primera pagina es la de posición 1
        _Sheet = _Writer.active
        self.__Sheet=_Sheet
        Header_f, Listado = [n.value for n in _Sheet[1] if n.value not in (None,'None','',' ')], []#_Sheet[1] : es la primera fila de la hoja
        maximo=len(Header_f)
        #se filtran las cabeceras
        cambio={'á':'a','é':'e','í':'i','ó':'o','ú':'u',' ':'_','/':'_','ñ':'n'}
        Header=["".join([cambio[y].upper() if y in cambio else y.upper() for y in x.lower().replace(' / ',' ')]) for x in Header_f]
        self.Header=Header
        for row in _Sheet.iter_rows():
            lista=[x.value for x in row[:maximo]]
            if lista==[None]*len(lista) or Header_f==lista:continue
            Listado.append(dict(zip(Header,map(filtro,lista))))
        if read_only:
            del _Sheet
            _Writer.close()
            return Listado
        Sheet =lambda idx,name:_Sheet.cell(row=2+idx, column=Header.index(name)+1)## Sheet(0,'QUIEBRE').value='hola'
        Writer=lambda:_Writer.save(self.excel)
        return Listado,Sheet,Writer

    def Xls2Dict(self,Excel):# Mejorar funcion
        ##  https://stackoverflow.com/questions/12250024/how-to-obtain-sheet-names-from-xls-files-without-loading-the-whole-file/12250416#12250416
        def filtro(value):
            value = value if value not in (None,'None','',' ') else ''
            value = value if not(isinstance(value, datetime)) else value.strftime("%Y-%m-%d")#"%d/%m/%Y")
            return str(value).replace("'", "")
        Writer = xlrd.open_workbook(filename=Excel, on_demand = True)
        all_sheets = Writer.sheets()
        Sheet = all_sheets[0]
        Header_f, Listado = list(n for n in Sheet.row_values(0) if n not in (None,'')), []
        maximo=len(Header_f)
        #se filtran las cabeceras
        cambio={'á':'a','é':'e','í':'i','ó':'o','ú':'u',' ':'_','/':'_','ñ':'n'}
        Header=["".join([cambio[y].upper() if y in cambio else y.upper() for y in x.lower().replace(' / ',' ')]) for x in Header_f]    
        for row in range(1,Sheet.nrows):
            lista=Sheet.row_values(row)[:maximo]
            if lista in ([None]*len(lista),['']*len(lista)) or Header_f==lista:continue
            Listado.append(dict(zip(Header,map(filtro,lista))))
        Writer.release_resources()
        del Writer
        return Listado

    def make_excel(self,resultado,dst,name):# Mejorar funcion
        """
        resultado: resultado del query
        dst: ruta de de destino
        name: nombre del archivo
        """
        #crea un valor de amdocs
        print('resultado',len(resultado))
        if len(resultado)==0:return
        if isinstance(resultado[0],dict):
            Header=[x for x in resultado[0]]
            body = [[x[y] for y in x] for x in resultado]
            resultado = [Header] + body
        wb=openpyxl.Workbook()#(write_only=True)
        sheet =wb.active
        sheet.title = "Hoja 1"
        filtro=lambda x: str(int(x))if isinstance(x,decimal.Decimal) else (x.strftime("%d/%m/%Y %H:%M:%S") if isinstance(x,datetime) else ( "" if x==None else x))
        rows = [list(map(filtro,x)) for x in resultado]
        cabecera = openpyxl.styles.PatternFill(start_color='005B9BD5', end_color='005B9BD5', fill_type='solid')
        par=openpyxl.styles.PatternFill(start_color='00DEEBF6', end_color='00DEEBF6', fill_type='solid')
        error=[]
        r=[]
        for row in rows:
            try:
                sheet.append(row)
            except:
                print(row)# error por caracteres especiales
                error.append(row)
        columnas = sheet.max_column
        filas = sheet.max_row
        for y in range(filas,filas+len(error)):
            for x in range(columnas):
                try:
                    sheet.cell(row=y+1,column=x+1).value=str(error[y-filas][x]).replace('\x03','').replace('\x13','')
                except:
                    r.append(error[y-filas][x])
                    print(error[y-filas][x])
        columnas = sheet.max_column
        filas = sheet.max_row
        ## cabeceras color
        for x in range(columnas):
            sheet.cell(row=1, column=x+1).font = openpyxl.styles.Font(bold=True, color=openpyxl.styles.colors.WHITE)
            sheet.cell(row=1, column=x+1).fill = cabecera
        ## datos color
        for y in range(1,filas):
            for x in range(columnas):
                if y%2==1:
                    break
                else:
                   sheet.cell(row=y+1, column=x+1).fill = par
        wb.save(filename = os.path.join(dst,name))
        del wb


class Threader(threading.Thread):#controller two button at the same time
    
    def __init__(self, *args, **kwargs):
        threading.Thread.__init__(self, *args, **kwargs)
        self.daemon = True
        self.killed = False
        self.flag="play"
        
    def start(self):
        self.__run_backup = self.run
        self.run = self.__run
        threading.Thread.start(self)
  
    def __run(self):
        sys.settrace(self.globaltrace) 
        self.__run_backup() # Fuerza al hilo a instalar nuestra traza
        self.run = self.__run_backup 
  
    def globaltrace(self, frame, event, arg):
        if event == 'call':
            return self.localtrace 
        else:
            return None
      
    def localtrace(self, frame, event, arg):
        if self.flag=="pausa":
            while True:
                if self.flag=="play":
                    break
                pya.sleep(0.1)
                if self.killed:
                    raise SystemExit()
        elif self.killed:
            if event == 'line':
                raise SystemExit()
        return self.localtrace
      
    def kill(self):
        self.killed = True
        self.join()
        
    def pause(self):
        self.flag="pausa"
        
    def play(self):
        self.flag="play"
          
if __name__=="__main__":
    pass
    